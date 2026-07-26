# -*- coding: utf-8 -*-
"""
database.py — низкоуровневый слой доступа к данным.

В V2 "базой данных" служат Excel-файлы в cloud_storage/. Этот модуль
даёт единый интерфейс чтения/записи строк как списков словарей,
а также умеет автоматически создавать файлы с нужной структурой,
если их нет (см. раздел 3 документации).
"""

import os
import shutil
import threading
from datetime import datetime, timedelta

import openpyxl
from openpyxl import Workbook

import config

# Глобальная блокировка на файловые операции (простая защита от гонок,
# так как несколько апдейтов Telegram могут обрабатываться параллельно)
_lock = threading.Lock()


# ------------------------------------------------------------------
#  Базовые операции чтения/записи листов
# ------------------------------------------------------------------

def read_sheet(path, sheet_name=None):
    """
    Читает лист Excel-файла и возвращает (headers, rows), где
    headers — список названий колонок, rows — список словарей.
    """
    if not os.path.exists(path):
        return [], []

    with _lock:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            wb.close()
            return [], []

        rows = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                value = raw_row[i] if i < len(raw_row) else None
                row_dict[header] = value
            rows.append(row_dict)

        wb.close()
        return headers, rows


def write_sheet(path, headers, rows, sheet_name=None):
    """
    Полностью перезаписывает лист заданными headers/rows.
    rows — список словарей header -> value.
    """
    with _lock:
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row)
            elif sheet_name:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb.active
                ws.delete_rows(1, ws.max_row)
        else:
            wb = Workbook()
            ws = wb.active
            if sheet_name:
                ws.title = sheet_name

        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        wb.save(path)


def append_row(path, row_dict, sheet_name=None):
    """Добавляет одну строку в конец листа, сохраняя существующие заголовки."""
    headers, rows = read_sheet(path, sheet_name)
    if not headers:
        headers = list(row_dict.keys())
    rows.append(row_dict)
    write_sheet(path, headers, rows, sheet_name)


def update_row(path, match_fn, updates, sheet_name=None):
    """
    Находит первую строку, для которой match_fn(row) == True,
    и обновляет её значениями из updates. Возвращает True, если строка найдена.
    """
    headers, rows = read_sheet(path, sheet_name)
    found = False
    for row in rows:
        if match_fn(row):
            row.update(updates)
            found = True
            break
    if found:
        for key in updates:
            if key not in headers:
                headers.append(key)
        write_sheet(path, headers, rows, sheet_name)
    return found


def find_row(path, match_fn, sheet_name=None):
    """Возвращает первую строку, удовлетворяющую match_fn, либо None."""
    _, rows = read_sheet(path, sheet_name)
    for row in rows:
        if match_fn(row):
            return row
    return None


def find_rows(path, match_fn, sheet_name=None):
    """Возвращает все строки, удовлетворяющие match_fn."""
    _, rows = read_sheet(path, sheet_name)
    return [row for row in rows if match_fn(row)]


def backup_file(path):
    """Копирует файл в cloud_storage/backups/ с меткой времени."""
    if not os.path.exists(path):
        return
    backups_dir = os.path.join(config.CLOUD_STORAGE_DIR, "backups")
    os.makedirs(backups_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = os.path.basename(path)
    dest = os.path.join(backups_dir, f"{stamp}_{name}")
    shutil.copy2(path, dest)


# ------------------------------------------------------------------
#  Создание файлов с нуля (раздел 3 документации)
# ------------------------------------------------------------------

def create_workbook_with_sheets(path, sheets):
    """
    sheets: dict sheet_name -> (headers, rows)
    Создаёт файл только если он не существует или пуст.
    """
    if os.path.exists(path) and os.path.getsize(path) > 0:
        return False  # уже существует, не трогаем

    wb = Workbook()
    first = True
    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return True


def ensure_system_config():
    """Создаёт system_config.xlsx с листами Компании/Приглашения/Администраторы/Настройки."""
    path = config.SYSTEM_CONFIG_FILE
    created = create_workbook_with_sheets(path, {
        "Компании": (
            ["company_id", "company_name", "is_active", "trial_ends_at"],
            [],
        ),
        "Приглашения": (
            ["ID", "Статус", "Компания"],
            [],
        ),
        "Администраторы": (
            ["admin_login", "admin_password", "company_id", "role"],
            [],
        ),
        "Настройки": (
            ["setting_key", "setting_value"],
            [],
        ),
    })
    return created


def ensure_company_files(company_dir):
    """
    Создаёт (при отсутствии) все Excel-файлы компании:
    пользователи.xlsx, партии.xlsx, учёт_времени.xlsx, брак.xlsx,
    конвейер_настройки.xlsx — см. раздел 3.2 документации.
    """
    os.makedirs(company_dir, exist_ok=True)

    users_path = os.path.join(company_dir, config.COMPANY_FILE_NAMES["users"])
    create_workbook_with_sheets(users_path, {
        "Лист1": (
            ["логин", "пароль", "имя", "роль", "процессы"],
            [[config.DEFAULT_ADMIN_LOGIN, config.DEFAULT_ADMIN_PASSWORD,
              "Макс Админ", config.ROLE_BOSS, ""]],
        )
    })

    batches_path = os.path.join(company_dir, config.COMPANY_FILE_NAMES["batches"])
    create_workbook_with_sheets(batches_path, {
        "Лист1": (
            ["id_партии", "название", "текущий_процесс", "статус",
             "создана", "прогноз_готовности"],
            [],
        )
    })

    time_path = os.path.join(company_dir, config.COMPANY_FILE_NAMES["time_tracking"])
    create_workbook_with_sheets(time_path, {
        "Лист1": (
            ["логин", "дата", "часы", "записал"],
            [],
        )
    })

    defects_path = os.path.join(company_dir, config.COMPANY_FILE_NAMES["defects"])
    create_workbook_with_sheets(defects_path, {
        "Лист1": (
            ["id", "логин", "процесс", "дата", "фото", "статус", "комментарий"],
            [],
        )
    })

    conveyor_path = os.path.join(company_dir, config.COMPANY_FILE_NAMES["conveyor"])
    create_workbook_with_sheets(conveyor_path, {
        "Лист1": (
            ["ID", "Название", "Ответственный", "Зависимость_от",
             "Обязательные", "Цена", "Норма_времени", "Фотоотчёт"],
            [
                [0, "Раскрой", "Максат", "", "", 12, 45, "Да"],
                [1, "Процесс 1", "", 0, 0, "", "", ""],
                [2, "Процесс 2", "Гулжамал", 0, 0, "", "", ""],
                [3, "Процесс 3 (опц.)", "Нуржамал", 0, "", "", "", ""],
                [4, "Сборка", "Нуржамал", "1,2,3", "1,2", "", "", ""],
            ],
        )
    })

    return {
        "users": users_path,
        "batches": batches_path,
        "time_tracking": time_path,
        "defects": defects_path,
        "conveyor": conveyor_path,
    }


def bootstrap_default_company():
    """
    При самом первом запуске бота: создаёт system_config.xlsx,
    компанию TelegramBot ERP и её файлы (раздел 2.4 документации).
    Возвращает dict с данными созданной компании, либо None,
    если компания уже существовала.
    """
    config.ensure_directories()
    ensure_system_config()

    companies_headers, companies_rows = read_sheet(
        config.SYSTEM_CONFIG_FILE, "Компании"
    )
    for row in companies_rows:
        if row.get("company_name") == config.DEFAULT_COMPANY_NAME:
            return None  # уже создана раньше

    company_id = generate_company_id()
    trial_ends = (datetime.now() + timedelta(days=365 * 10)).strftime("%Y-%m-%d")

    append_row(
        config.SYSTEM_CONFIG_FILE,
        {
            "company_id": company_id,
            "company_name": config.DEFAULT_COMPANY_NAME,
            "is_active": "да",
            "trial_ends_at": trial_ends,
        },
        sheet_name="Компании",
    )

    append_row(
        config.SYSTEM_CONFIG_FILE,
        {
            "admin_login": config.DEFAULT_ADMIN_LOGIN,
            "admin_password": config.DEFAULT_ADMIN_PASSWORD,
            "company_id": company_id,
            "role": config.ROLE_BOSS,
        },
        sheet_name="Администраторы",
    )

    company_dir = os.path.join(config.CLOUD_STORAGE_DIR, company_id)
    ensure_company_files(company_dir)

    return {
        "company_id": company_id,
        "company_name": config.DEFAULT_COMPANY_NAME,
        "admin_login": config.DEFAULT_ADMIN_LOGIN,
        "admin_password": config.DEFAULT_ADMIN_PASSWORD,
    }


def generate_company_id(length=8):
    """Генерирует случайный ID компании из заглавных букв и цифр."""
    import random
    import string
    alphabet = string.ascii_uppercase + string.digits
    while True:
        candidate = "".join(random.choices(alphabet, k=length))
        existing = find_row(
            config.SYSTEM_CONFIG_FILE,
            lambda r: r.get("company_id") == candidate,
            sheet_name="Компании",
        )
        if not existing:
            return candidate


def company_dir_path(company_id):
    return os.path.join(config.CLOUD_STORAGE_DIR, company_id)
